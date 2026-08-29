"""资料来源文件生成（source_gen）：列补充、隐私与格式校验。"""
import json

from app.source_gen import generate

SAMPLE = "sample/课程表样例_脱敏.xlsx"


def test_generate_adds_three_columns(tmp_path):
    r = generate(SAMPLE, "2026/8/28", out_dir=tmp_path)
    assert r["rows"] > 0 and r["cols"] > 0

    import openpyxl
    ws = openpyxl.load_workbook(tmp_path / "资料来源_按条件查询上课情况.xlsx").active
    header = [c.value for c in ws[1]]
    assert header[-3:] == ["课程类别", "课程归属", "余量"]
    assert ws.max_row == r["rows"] + 1  # 表头 + 数据
    assert "教师联系电话" not in header  # 隐私红线兜底
    assert all(not str(h).startswith("Unnamed") for h in header if h)  # 无名空列已剔除

    meta = json.loads((tmp_path / "source_meta.json").read_text(encoding="utf-8"))
    assert meta["updated_at"] == "2026/8/28"
    assert meta["rows"] == r["rows"] and meta["cols"] == r["cols"]
    assert meta["size_bytes"] > 0
    # 余量 = 教学班人数 - 选课人数，抽查首行可核算
    import pandas as pd
    df = pd.read_excel(tmp_path / "资料来源_按条件查询上课情况.xlsx")
    row0 = df.iloc[0]
    assert row0["余量"] == round(row0["教学班人数"] - row0["选课人数"], 2)


def test_generate_rejects_wrong_format(tmp_path):
    import pandas as pd
    bad = tmp_path / "bad.xlsx"
    pd.DataFrame({"随便": [1]}).to_excel(bad, index=False)
    try:
        generate(bad, "2026/8/28", out_dir=tmp_path)
        assert False, "应当抛出 ValueError"
    except ValueError:
        pass
